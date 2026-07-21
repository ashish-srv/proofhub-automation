"""
combine_reconcile.py

Combines:
  - proofhub_tasks.csv          (from proofhub_single_project_fetch.py)
  - Quotation_Required_Data.csv (from your existing quotations script)
  - ProofHub_ProjectName_ClientName_Mapping.xlsx
        tab: ProofHub_Client_Mapping   -> columns: 'ProofHub Project Name', 'Client Name'
        tab: Quotation_Client_Mapping  -> columns: 'Quotation Client Name', 'Client Name'

Does:
  1. Loads all three inputs.
  2. Maps ProofHub 'Project' -> 'Client Name' via ProofHub_Client_Mapping.
  3. Maps quotation 'client_name' -> 'Client Name' via Quotation_Client_Mapping.
  4. Any ProofHub Project / quotation client_name NOT found in the mapping is
     auto-appended to the relevant tab (Client Name left blank) and the excel
     file is saved back in place. Those rows are skipped this run until you
     fill in Client Name and re-run.
  5. Groups quotation data to one row per quotation (Client Name, quotation_id,
     quotation_client_name, quotation_start_date, quotation_end_date,
     quotation_status, quotation_location, planned_creatives = summed
     Creatives across its format rows).
  6. Keeps EVERY ProofHub task row exactly as fetched (no workflow/stage
     filtering, no date filtering) and left-joins the matching quotation's
     planned_creatives onto each task, where the task's Client Name matches
     and its Created At falls within that quotation's start_date-end_date.
     A task can match more than one quotation if a client has overlapping
     quotation periods -> it will appear once per match (flagged in console).
     A task with no matching quotation still appears, with blank quotation columns.
  7. Writes two outputs:
        quotation_format_detail.csv       -> planned creatives by format (unchanged detail)
        proofhub_tasks_with_quotation.csv -> full task-level data + joined quotation
                                              info, ready for your own pivot/date filter
"""

import os
import sys
import pandas as pd
from openpyxl import load_workbook

# ─────────────────────────────────────────────
# CONFIG — filenames (same folder as this script)
# ─────────────────────────────────────────────
PROOFHUB_CSV   = "proofhub_tasks.csv"
QUOTATION_CSV  = "Quotation_Required_Data.csv"
MAPPING_EXCEL  = "ProofHub_ProjectName_ClientName_Mapping.xlsx"

PROOFHUB_SHEET   = "ProofHub_Client_Mapping"
QUOTATION_SHEET  = "Quotation_Client_Mapping"

PROOFHUB_COL_PROJECT = "ProofHub Project Name"
PROOFHUB_COL_CLIENT  = "Client Name"
QUOTATION_COL_CLIENT_RAW = "Quotation Client Name"
QUOTATION_COL_CLIENT     = "Client Name"

REQUIRED_QUOTATION_FIELDS = [
    "quotation_id", "client_name", "location", "start_date", "end_date",
    "status", "Department", "Sub Department", "Format", "Creatives"
]


# ─────────────────────────────────────────────
# STEP 0 — sanity checks
# ─────────────────────────────────────────────
def require_file(path):
    if not os.path.exists(path):
        print(f"❌ Required file not found: {path}")
        sys.exit(1)


require_file(PROOFHUB_CSV)
require_file(QUOTATION_CSV)
require_file(MAPPING_EXCEL)


# ─────────────────────────────────────────────
# STEP 1 — load inputs
# ─────────────────────────────────────────────
proofhub_df  = pd.read_csv(PROOFHUB_CSV, dtype=str, keep_default_na=False)
quotation_df = pd.read_csv(QUOTATION_CSV, dtype=str, keep_default_na=False)

WORKFLOW_FILTER = "creative creation"  # case-insensitive match
before_count = len(proofhub_df)
proofhub_df = proofhub_df[proofhub_df["Workflow"].str.strip().str.casefold() == WORKFLOW_FILTER].copy()
print(f"ℹ Filtered to Workflow = 'Creative Creation': {len(proofhub_df)} of {before_count} tasks kept")

missing_cols = [c for c in REQUIRED_QUOTATION_FIELDS if c not in quotation_df.columns]
if missing_cols:
    print(f"❌ Quotation_Required_Data.csv is missing expected columns: {missing_cols}")
    sys.exit(1)

quotation_df = quotation_df[REQUIRED_QUOTATION_FIELDS].copy()

proofhub_map_df  = pd.read_excel(MAPPING_EXCEL, sheet_name=PROOFHUB_SHEET, dtype=str)
quotation_map_df = pd.read_excel(MAPPING_EXCEL, sheet_name=QUOTATION_SHEET, dtype=str)

proofhub_map_df  = proofhub_map_df.fillna("")
quotation_map_df = quotation_map_df.fillna("")


# ─────────────────────────────────────────────
# STEP 2 — map ProofHub Project -> Client Name
# ─────────────────────────────────────────────
proofhub_lookup = dict(zip(
    proofhub_map_df[PROOFHUB_COL_PROJECT].str.strip(),
    proofhub_map_df[PROOFHUB_COL_CLIENT].str.strip()
))

proofhub_df["Client Name"] = proofhub_df["Project"].str.strip().map(proofhub_lookup)

unmapped_projects = sorted(
    proofhub_df.loc[proofhub_df["Client Name"].isna(), "Project"].dropna().unique().tolist()
)
blank_mapped_projects = sorted(set(
    proofhub_df.loc[proofhub_df["Client Name"] == "", "Project"].dropna().unique().tolist()
))
unmapped_projects = sorted(set(unmapped_projects) | set(blank_mapped_projects))

# ─────────────────────────────────────────────
# STEP 3 — map quotation client_name -> Client Name
# ─────────────────────────────────────────────
quotation_lookup = dict(zip(
    quotation_map_df[QUOTATION_COL_CLIENT_RAW].str.strip(),
    quotation_map_df[QUOTATION_COL_CLIENT].str.strip()
))

quotation_df["Client Name"] = quotation_df["client_name"].str.strip().map(quotation_lookup)

unmapped_quotation_clients = sorted(
    quotation_df.loc[quotation_df["Client Name"].isna(), "client_name"].dropna().unique().tolist()
)
blank_mapped_quotation_clients = sorted(set(
    quotation_df.loc[quotation_df["Client Name"] == "", "client_name"].dropna().unique().tolist()
))
unmapped_quotation_clients = sorted(set(unmapped_quotation_clients) | set(blank_mapped_quotation_clients))


# ─────────────────────────────────────────────
# STEP 4 — auto-append new/unmapped entries to the excel
# ─────────────────────────────────────────────
def append_new_mapping_rows(path, sheet_name, col_a_name, col_b_name, new_values):
    """Append rows with col_a filled and col_b (Client Name) left blank."""
    if not new_values:
        return

    wb = load_workbook(path)
    ws = wb[sheet_name]

    headers = [c.value for c in ws[1]]
    try:
        col_a_idx = headers.index(col_a_name) + 1
        col_b_idx = headers.index(col_b_name) + 1
    except ValueError:
        print(f"❌ Could not find expected headers in '{sheet_name}': {col_a_name} / {col_b_name}")
        return

    existing_values = set()
    for row in ws.iter_rows(min_row=2, values_only=False):
        val = row[col_a_idx - 1].value
        if val:
            existing_values.add(str(val).strip())

    next_row = ws.max_row + 1
    added = 0
    for val in new_values:
        if val in existing_values:
            continue
        ws.cell(row=next_row, column=col_a_idx, value=val)
        ws.cell(row=next_row, column=col_b_idx, value="")  # left blank on purpose
        next_row += 1
        added += 1

    if added:
        wb.save(path)
        print(f"📝 Added {added} new unmapped value(s) to '{sheet_name}' (Client Name left blank — fill in and re-run).")


append_new_mapping_rows(
    MAPPING_EXCEL, PROOFHUB_SHEET, PROOFHUB_COL_PROJECT, PROOFHUB_COL_CLIENT, unmapped_projects
)
append_new_mapping_rows(
    MAPPING_EXCEL, QUOTATION_SHEET, QUOTATION_COL_CLIENT_RAW, QUOTATION_COL_CLIENT, unmapped_quotation_clients
)

if unmapped_projects:
    print(f"⚠ {len(unmapped_projects)} ProofHub project(s) had no Client Name mapping: {unmapped_projects}")
if unmapped_quotation_clients:
    print(f"⚠ {len(unmapped_quotation_clients)} quotation client_name(s) had no Client Name mapping: {unmapped_quotation_clients}")

# Note: unmapped ProofHub tasks are still kept in the task-level output
# (Client Name blank) since step 6 keeps every task row regardless.
# Unmapped quotations are dropped from the quotation summary since they
# can't be matched to any client.
quotation_df = quotation_df[quotation_df["Client Name"].notna() & (quotation_df["Client Name"] != "")]


# ─────────────────────────────────────────────
# STEP 5 — group quotation data to one row per quotation
# ─────────────────────────────────────────────
quotation_df["Creatives"] = pd.to_numeric(quotation_df["Creatives"], errors="coerce").fillna(0)

# Detail table (planned, by department/format) — kept as-is, Client Name added
quotation_format_detail = quotation_df.copy()

quotation_summary = (
    quotation_df
    .groupby(["Client Name", "quotation_id", "client_name", "start_date", "end_date", "status", "location"], dropna=False)
    ["Creatives"].sum()
    .reset_index()
    .rename(columns={
        "Creatives": "planned_creatives",
        "client_name": "quotation_client_name",
        "start_date": "quotation_start_date",
        "end_date": "quotation_end_date",
        "status": "quotation_status",
        "location": "quotation_location",
    })
)

quotation_summary["quotation_start_date_parsed"] = pd.to_datetime(
    quotation_summary["quotation_start_date"], errors="coerce"
).dt.date
quotation_summary["quotation_end_date_parsed"] = pd.to_datetime(
    quotation_summary["quotation_end_date"], errors="coerce"
).dt.date


# ─────────────────────────────────────────────
# STEP 6 — keep every ProofHub task as-is, join matching quotation(s)
# ─────────────────────────────────────────────
proofhub_df["Created At (date)"] = pd.to_datetime(
    proofhub_df["Created At"], errors="coerce", utc=True
).dt.date

# left join on Client Name (many-to-many: a client can have several quotations)
joined = proofhub_df.merge(
    quotation_summary,
    on="Client Name",
    how="left",
    suffixes=("", "")
)

in_window = (
    joined["quotation_start_date_parsed"].notna()
    & joined["quotation_end_date_parsed"].notna()
    & (joined["Created At (date)"] >= joined["quotation_start_date_parsed"])
    & (joined["Created At (date)"] <= joined["quotation_end_date_parsed"])
)

matched_in_window = joined[in_window].drop(columns=["quotation_start_date_parsed", "quotation_end_date_parsed"])

# tasks with NO quotation match at all (either the client has no quotations,
# or it has quotations but none cover this task's Created At date) — keep
# them as single rows with blank quotation columns, don't drop them.
matched_task_ids = set(matched_in_window["Task ID"])
quotation_cols = [c for c in matched_in_window.columns if c not in proofhub_df.columns]

unmatched = proofhub_df[~proofhub_df["Task ID"].isin(matched_task_ids)].copy()
for c in quotation_cols:
    unmatched[c] = pd.NA

matched = pd.concat([matched_in_window, unmatched], ignore_index=True)

# ── collapse overlapping-quotation duplicates into one row per task ──
dupe_task_ids = matched.loc[
    matched.duplicated("Task ID", keep=False) & matched["quotation_id"].notna(), "Task ID"
].unique()
if len(dupe_task_ids):
    print(f"⚠ {len(dupe_task_ids)} task(s) matched more than one overlapping quotation period — "
          f"combining into a single row per task (planned_creatives summed, quotation_id joined).")

def join_unique(series):
    vals = sorted(set(str(x) for x in series.dropna() if str(x) != ""))
    return "; ".join(vals) if vals else pd.NA

def sum_or_blank(series):
    non_null = pd.to_numeric(series, errors="coerce").dropna()
    return non_null.sum() if len(non_null) else pd.NA

proofhub_cols = [c for c in proofhub_df.columns if c != "Created At (date)"] + ["Created At (date)"]
quotation_join_cols = ["quotation_id", "quotation_client_name", "quotation_status", "quotation_location"]

agg_map = {c: "first" for c in proofhub_cols if c != "Task ID"}
for c in quotation_join_cols:
    agg_map[c] = join_unique
agg_map["planned_creatives"] = sum_or_blank
agg_map["quotation_start_date"] = "min"
agg_map["quotation_end_date"] = "max"

result = matched.groupby("Task ID", as_index=False).agg(agg_map)

# add Month + Year directly on the task-level output too, not just the rollup
MONTH_LABELS = {
    1: "01. January", 2: "02. February", 3: "03. March", 4: "04. April",
    5: "05. May", 6: "06. June", 7: "07. July", 8: "08. August",
    9: "09. September", 10: "10. October", 11: "11. November", 12: "12. December",
}
_created_dt = pd.to_datetime(result["Created At"], errors="coerce", utc=True)
result["Year"] = _created_dt.dt.year
result["Month"] = _created_dt.dt.month.map(MONTH_LABELS)


# ─────────────────────────────────────────────
# STEP 6b — monthly Client + Stage rollup (fewer rows, for Zoho import)
# ─────────────────────────────────────────────
rollup_source = result.copy()
# Year and Month already exist on result now, reused here

def first_non_null(series):
    non_null = series.dropna()
    return non_null.iloc[0] if len(non_null) else pd.NA

monthly_rollup = (
    rollup_source
    .groupby(["Client Name", "Project", "Year", "Month", "Workflow", "Stage"], dropna=False)
    .agg(
        task_count=("Task ID", "count"),
        quotation_id=("quotation_id", join_unique),
        quotation_client_name=("quotation_client_name", join_unique),
        quotation_status=("quotation_status", join_unique),
        quotation_start_date=("quotation_start_date", first_non_null),
        quotation_end_date=("quotation_end_date", first_non_null),
        planned_creatives=("planned_creatives", "max"),
    )
    .reset_index()
)

# Month labels start with "01.", "02." etc., so sorting the string sorts
# chronologically within a year — no separate date parsing needed.
monthly_rollup = monthly_rollup.sort_values(["Client Name", "Year", "Month"])


# ─────────────────────────────────────────────
# STEP 7 — save outputs
# ─────────────────────────────────────────────
OUTPUT_EXCEL = "Creatives_Dashboard_Data.xlsx"
quotation_format_detail_out = quotation_format_detail.rename(columns={"client_name": "Quotation_Client_Name"})

with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
    quotation_format_detail_out.to_excel(writer, sheet_name="Quotation_Format_Detail", index=False)
    result.to_excel(writer, sheet_name="ProofHub_Tasks_With_Quotation", index=False)
    monthly_rollup.to_excel(writer, sheet_name="Monthly_Client_Status_Rollup", index=False)

print(f"\n✅ {OUTPUT_EXCEL} written:")
print(f"   - Quotation_Format_Detail: {len(quotation_format_detail_out)} rows")
print(f"   - ProofHub_Tasks_With_Quotation: {len(result)} rows ({len(proofhub_df)} raw ProofHub tasks in)")
print(f"   - Monthly_Client_Status_Rollup: {len(monthly_rollup)} rows")
print("\n" + "=" * 50)
print("DONE")
print("=" * 50)
